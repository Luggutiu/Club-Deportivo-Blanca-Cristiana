# ========================================
# Proyecto desarrollado exclusivamente para:
# Club Deportivo Blanca Cristiana
# Desarrollador: Luis Gutierrez
# Sitio: https://club-deportivo-blanca-cristiana.onrender.com
# Email: clubdeportivoblancacristiana@gmail.com
# Año: 2025
# Todos los derechos reservados
# ========================================
from fastapi import FastAPI, Request, Depends, Form, UploadFile, File, HTTPException, Query
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, FileResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.exceptions import RequestValidationError
from starlette.middleware.sessions import SessionMiddleware
from starlette.status import HTTP_303_SEE_OTHER
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Alignment, PatternFill
from fastapi.responses import StreamingResponse
from io import BytesIO
from datetime import datetime
from fastapi import status
from fastapi.responses import RedirectResponse
import os
from sqlalchemy.orm import Session
from io import BytesIO
from openpyxl import Workbook

from app.database import get_db
from app.models import Suscriptor
from app.routes.auth import check_admin_logged
from app.routes import admin_info

# ------------------------- Utilidades estándar -------------------------
from datetime import datetime, date
import os
import shutil

# ------------------------- Base de Datos y Modelos -------------------------
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError
from app.database import get_db
from app.models import Post, Horario, SeccionInformativa, Suscriptor

# ------------------------- Utilidades del proyecto -------------------------
from app.routes.embedder import generar_embed
from app.routes.auth import check_admin_logged
from app.utils.email_utils import enviar_correo_bienvenida, notificar_admin_suscripcion

# ------------------------- Rutas personalizadas -------------------------
from app.routes import (
    like,
    auth,
    admin,
    posts,
    dev,
    auth_google,
    healthcheck
)
from app.routes.suscripcion import router as suscripcion_router


# Inicialización
app = FastAPI()

app.mount("/static", StaticFiles(directory="app/static"), name="static")
templates = Jinja2Templates(directory="app/templates")

# Montar rutas
app.include_router(auth.router)
app.include_router(like.router)
app.include_router(healthcheck.router)
app.include_router(auth_google.router)
app.include_router(suscripcion_router)
app.include_router(admin_info.router)
app.include_router(admin.router)
app.include_router(posts.router)
app.include_router(dev.router)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

@app.get("/sitemap.xml", include_in_schema=False)
def sitemap(request: Request, db: Session = Depends(get_db)):
    base_url = str(request.base_url).rstrip("/")

    urls = [{"loc": f"{base_url}/", "priority": "1.0"}]

    # 🔄 Agregar secciones informativas dinámicamente desde la base
    secciones = db.query(SeccionInformativa).all()
    for seccion in secciones:
        urls.append({"loc": f"{base_url}/{seccion.slug}"})

    # 📰 Agregar los posts dinámicamente
    posts = db.query(Post).order_by(Post.id.desc()).all()
    for post in posts:
        urls.append({
            "loc": f"{base_url}/post/{post.id}",
            "lastmod": post.fecha_creacion.strftime("%Y-%m-%d") if post.fecha_creacion else None
        })

    # 🛠️ Generar XML
    xml_content = '<?xml version="1.0" encoding="UTF-8"?>\n'
    xml_content += '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n'

    for url in urls:
        xml_content += "  <url>\n"
        xml_content += f"    <loc>{url['loc']}</loc>\n"
        if "priority" in url:
            xml_content += f"    <priority>{url['priority']}</priority>\n"
        if "lastmod" in url and url["lastmod"]:
            xml_content += f"    <lastmod>{url['lastmod']}</lastmod>\n"
        xml_content += "  </url>\n"

    xml_content += "</urlset>"

    return Response(content=xml_content, media_type="application/xml")

@app.get("/robots.txt", include_in_schema=False)
def robots():
    robots_path = os.path.join(BASE_DIR, "static", "robots.txt")
    return FileResponse(robots_path, media_type="text/plain")

app.add_middleware(
    SessionMiddleware,
    secret_key=os.getenv("SECRET_KEY", "2025*")
) # Reemplaza por variable de entorno


@app.post("/registrar-google-suscriptor")
async def registrar_google_suscriptor(
    request: Request,
    correo: str = Form(...),
    nombre_completo: str = Form(...),
    tipo_documento: str = Form(...),
    numero_documento: str = Form(...),
    celular: str = Form(...),
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    try:
        # Verificar si ya está registrado
        existe = db.query(Suscriptor).filter(Suscriptor.numero_documento == numero_documento).first()
        if existe:
            return templates.TemplateResponse("error.html", {
                "request": request,
                "error_message": "Este número de documento ya está registrado."
            })

        # Crear el suscriptor
        nuevo = Suscriptor(
            nombre_completo=nombre_completo,
            correo=correo,
            tipo_documento=tipo_documento,
            numero_documento=numero_documento,
            celular=celular
        )
        db.add(nuevo)
        db.commit()

        # Leer el archivo en memoria (sin guardar en disco)
        contenido_bytes = await archivo.read()
        extension = archivo.filename.split('.')[-1].lower()
        filename = f"{numero_documento}.{extension}"

        # Enviar correos
        await enviar_correo_bienvenida(nombre_completo, correo)
        await notificar_admin_suscripcion(
            nombre=nombre_completo,
            correo=correo,
            tipo=tipo_documento,
            documento=numero_documento,
            celular=celular,
            archivo_path=None,
            archivo_bytes=contenido_bytes,
            archivo_nombre=filename
        )

        return RedirectResponse(url="/confirmacion-suscripcion", status_code=303)

    except Exception as e:
        print("❌ Error al registrar vía Google:", e)
        return templates.TemplateResponse("error.html", {
            "request": request,
            "error_message": "Ocurrió un error al procesar tu registro."
        })







@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request, exc):
    return JSONResponse(
        status_code=422,
        content={"detail": exc.errors(), "body": exc.body})

# --------------------- Rutas Públicas ---------------------

@app.get("/", response_class=HTMLResponse)
async def home(request: Request, db: Session = Depends(get_db)):
    try:
        posts = db.query(Post).all()
        horarios = db.query(Horario).filter(Horario.publicado == True).all()

        print("HORARIOS ENCONTRADOS:")
        for h in horarios:
            print(f"{h.dia} - {h.hora_inicio} a {h.hora_fin}")

        publicaciones = posts + horarios
        publicaciones.sort(key=lambda x: getattr(x, 'fecha_creacion', None) or x.id, reverse=True)

        return templates.TemplateResponse("index.html", {
            "request": request,
            "publicaciones": publicaciones,
        })
    except Exception as e:
        return templates.TemplateResponse("index.html", {
        "request": request,
        "publicaciones": publicaciones
    })
        

@app.get("/terminos-condiciones", response_class=HTMLResponse)
async def terminos_condiciones(request: Request):
    return templates.TemplateResponse("terminos_condiciones.html", {"request": request})

@app.get("/derechos-reservados", response_class=HTMLResponse)
async def derechos_reservados(request: Request):
    return templates.TemplateResponse("derechos_reservados.html", {
        "request": request,
        "current_year": datetime.now().year
    })

@app.get("/politica-privacidad", response_class=HTMLResponse)
def politica_privacidad(request: Request):
    return templates.TemplateResponse("politica_privacidad.html", {"request": request})

@app.get("/condiciones-servicio", response_class=HTMLResponse)
def condiciones_servicio(request: Request):
    return templates.TemplateResponse("condiciones_servicio.html", {"request": request})

@app.get("/contacto", response_class=HTMLResponse)
def contacto(request: Request, db=Depends(get_db)):
    seccion = db.query(SeccionInformativa).filter(SeccionInformativa.titulo == "contacto").first()
    secciones = db.query(SeccionInformativa).all()  # ✅ Para el menú lateral

    return templates.TemplateResponse("contacto.html", {
        "request": request,
        "contenido": seccion,
        "secciones": secciones  # ✅ Asegura que el menú funcione
    })

# --------------------- Suscripción clásica ---------------------

from fastapi.responses import JSONResponse

@app.get("/suscribirse", response_class=HTMLResponse)
async def mostrar_formulario_suscripcion(
    request: Request,
    success: str = "",
    error: str = "",
    nombre_completo: str = "",
    correo: str = "",
    tipo_documento: str = "",
    numero_documento: str = "",
    celular: str = "",
    acepto: bool = False,
    db: Session = Depends(get_db)  # 👈 se añade para obtener las secciones
):
    secciones = db.query(SeccionInformativa).all()
    return templates.TemplateResponse("suscribirse.html", {
        "request": request,
        "success": success,
        "error": error,
        "nombre_completo": nombre_completo,
        "correo": correo,
        "tipo_documento": tipo_documento,
        "numero_documento": numero_documento,
        "celular": celular,
        "acepto": acepto,
        "secciones": secciones  # 👈 añadido
    })


@app.get("/confirmacion-suscripcion", response_class=HTMLResponse)
def confirmacion_suscripcion(request: Request, db: Session = Depends(get_db)):
    nombre = request.session.get("nombre_completo")
    correo = request.session.get("correo")
    secciones = db.query(SeccionInformativa).all()
    return templates.TemplateResponse("confirmacion_suscripcion.html", {
        "request": request,
        "nombre": nombre,
        "correo": correo,
        "secciones": secciones
    })



@app.post("/suscribirse")
async def procesar_suscripcion(
    request: Request,
    nombre_completo: str = Form(...),
    correo: str = Form(...),
    tipo_documento: str = Form(...),
    numero_documento: str = Form(...),
    celular: str = Form(...),
    archivo: UploadFile = File(None),  # ← CORRECTO
    acepto: bool = Form(...),
    db: Session = Depends(get_db),
):
    try:
        print(">> Procesando suscripción")

        if not acepto:
            return JSONResponse(
                status_code=400,
                content={"error": "Debes aceptar los términos y condiciones."}
            )

        existe = db.query(Suscriptor).filter(Suscriptor.numero_documento == numero_documento).first()
        if existe:
            return JSONResponse(
                status_code=400,
                content={"error": "documento_existente"}
            )

        archivo_path = None
        if archivo:
            carpeta_destino = "app/static/archivos"
            os.makedirs(carpeta_destino, exist_ok=True)
            archivo_path = f"{carpeta_destino}/{numero_documento}_{archivo.filename}"
            with open(archivo_path, "wb") as buffer:
                shutil.copyfileobj(archivo.file, buffer)

        nuevo = Suscriptor(
            nombre_completo=nombre_completo,
            correo=correo,
            tipo_documento=tipo_documento,
            numero_documento=numero_documento,
            celular=celular,
            
        )
        db.add(nuevo)
        db.commit()
        
        print("DEBUG >> Datos recibidos:")
        print("Nombre:", nombre_completo)
        print("Correo:", correo)
        print("Tipo documento:", tipo_documento)
        print("Número documento:", numero_documento)
        print("Celular:", celular)

        await enviar_correo_bienvenida(nombre_completo, correo)
        await notificar_admin_suscripcion(nombre_completo, correo, tipo_documento, numero_documento, celular, archivo_path)

        return JSONResponse(
            status_code=200,
            content={"mensaje": "¡Gracias por unirte al club! Te hemos enviado un correo de bienvenida."}
        )

    except Exception as e:
        print("❌ Error al procesar suscripción:", str(e))
        return JSONResponse(
            status_code=500,
            content={"error": "registro_fallido"}
        )

@app.post("/guardar-suscriptor", response_class=HTMLResponse)
async def guardar_suscriptor(
    request: Request,
    tipo_documento: str = Form(...),
    numero_documento: str = Form(...),
    nombre_completo: str = Form(...),
    celular: str = Form(...),
    correo: str = Form(...),
    db: Session = Depends(get_db)
):
    nuevo = Suscriptor(
        tipo_documento=tipo_documento,
        numero_documento=numero_documento,
        nombre=nombre_completo,
        nombre_completo=nombre_completo,
        celular=celular,
        correo=correo
    )
    try:
        db.add(nuevo)
        db.commit()
        await enviar_correo_bienvenida(destinatario=correo, nombre=nombre_completo)
        await notificar_admin_suscripcion(nombre=nombre_completo, correo=correo, documento=numero_documento, tipo=tipo_documento, celular=celular)
        return templates.TemplateResponse("confirmacion_suscripcion.html", {
            "request": request,
            "nombre": nombre_completo,
            "correo": correo
        })
    except IntegrityError:
        db.rollback()
        return templates.TemplateResponse("error.html", {
            "request": request,
            "error_message": "Este correo o documento ya está registrado."
        }, status_code=400)

@app.get("/formulario-suscriptor", response_class=HTMLResponse)
def formulario_suscriptor(request: Request, correo: str = Query(None), nombre: str = Query(None), db: Session = Depends(get_db)):
    secciones = db.query(SeccionInformativa).all()
    return templates.TemplateResponse("formulario_suscriptor.html", {
        "request": request,
        "correo": correo,
        "nombre": nombre,
        "secciones": secciones
    })

# --------------------- Panel de Administración ---------------------

@app.get("/admin", response_class=HTMLResponse)
def admin_panel(request: Request, db: Session = Depends(get_db)):
    if not check_admin_logged(request):
        return RedirectResponse(url="/login", status_code=302)
    publicaciones = db.query(Post).all()
    horarios = db.query(Horario).filter(Horario.publicado == True).all()
    return templates.TemplateResponse("admin.html", {
        "request": request,
        "publicaciones": publicaciones,
        "horarios": horarios
    })

@app.get("/admin/publicar-post", response_class=HTMLResponse)
def formulario_publicar_post(request: Request):
    if not check_admin_logged(request):
        return RedirectResponse(url="/login", status_code=302)
    return templates.TemplateResponse("publicar_post.html", {"request": request})

@app.get("/test-embed", response_class=HTMLResponse)
def test_embed(request: Request):
    return templates.TemplateResponse("test_embed.html", {"request": request})

@app.get("/admin/gestionar-horarios", response_class=HTMLResponse)
def mostrar_formulario_horario(request: Request, db=Depends(get_db)):
    horarios = db.query(Horario).order_by(Horario.dia).all()

    print("HORARIOS EN ADMIN:")
    for h in horarios:
        print(f"- {h.dia} | {h.hora_inicio} - {h.hora_fin} | {h.actividad} | Publicado: {h.publicado}")

    return templates.TemplateResponse("gestionar_horarios.html", {
        "request": request,
        "horarios": horarios
    })

@app.get("/admin/editar-horario/{horario_id}", response_class=HTMLResponse)
def mostrar_formulario_edicion(request: Request, horario_id: int, db=Depends(get_db)):
    horario = db.query(Horario).filter(Horario.id == horario_id).first()
    if not horario:
        return HTMLResponse("Horario no encontrado", status_code=404)
    return templates.TemplateResponse("editar_horario.html", {
        "request": request,
        "horario": horario
    })

@app.post("/admin/guardar-horario")
def guardar_horario(
    request: Request,
    dia: str = Form(...),
    hora_inicio: str = Form(...),
    hora_fin: str = Form(...),
    actividad: str = Form(...),
    db=Depends(get_db)
):
    nuevo_horario = Horario(dia=dia, hora_inicio=hora_inicio, hora_fin=hora_fin, actividad=actividad, publicado=True)
    db.add(nuevo_horario)
    db.commit()
    return RedirectResponse(url="/admin/gestionar-horarios", status_code=303)


from fastapi import Request
from fastapi.responses import RedirectResponse
from app.database import SessionLocal
from app import models

@app.post("/admin/publicar-horario/{horario_id}")
def publicar_horario(horario_id: int, request: Request):
    db = SessionLocal()
    try:
        horario = db.query(models.Horario).filter(models.Horario.id == horario_id).first()
        if horario:
            horario.publicado = True
            db.commit()
    finally:
        db.close()
    
    return RedirectResponse(url="/admin/gestionar-horarios", status_code=303)

@app.post("/admin/publicar-post")
def guardar_post(request: Request, titulo: str = Form(...), texto: str = Form(None), imagen_url: str = Form(None), db=Depends(get_db)):
    nuevo_post = Post(titulo=titulo, texto=texto, imagen_url=imagen_url)
    db.add(nuevo_post)
    db.commit()
    return RedirectResponse(url="/admin", status_code=303)

@app.post("/admin/eliminar-post/{post_id}")
async def eliminar_post(post_id: int, db=Depends(get_db)):
    post = db.query(Post).filter(Post.id == post_id).first()
    if not post:
        raise HTTPException(status_code=404, detail="Publicación no encontrada")
    db.delete(post)
    db.commit()
    return RedirectResponse(url="/admin", status_code=HTTP_303_SEE_OTHER)

@app.get("/admin/editar-post/{post_id}", response_class=HTMLResponse)
def mostrar_formulario_editar_post(request: Request, post_id: int, db=Depends(get_db)):
    if not check_admin_logged(request):
        return RedirectResponse(url="/login", status_code=302)
    post = db.query(Post).filter(Post.id == post_id).first()
    if not post:
        return HTMLResponse("Post no encontrado", status_code=404)
    return templates.TemplateResponse("editar_post.html", {"request": request, "post": post})

@app.post("/admin/editar-post/{post_id}")
def editar_post(request: Request, post_id: int, titulo: str = Form(...), texto: str = Form(None), imagen_url: str = Form(None), db=Depends(get_db)):
    post = db.query(Post).filter(Post.id == post_id).first()
    if not post:
        return HTMLResponse("Post no encontrado", status_code=404)
    post.titulo = titulo
    post.texto = texto
    post.imagen_url = imagen_url
    db.commit()
    return RedirectResponse(url="/admin", status_code=303)

@app.post("/admin/editar-horario/{horario_id}")
def actualizar_horario(
    request: Request,
    horario_id: int,
    dia: str = Form(...),
    hora_inicio: str = Form(...),
    hora_fin: str = Form(...),
    actividad: str = Form(...),
    db=Depends(get_db)
):
    horario = db.query(Horario).filter(Horario.id == horario_id).first()
    if not horario:
        return HTMLResponse("Horario no encontrado", status_code=404)
    horario.dia = dia
    horario.hora_inicio = hora_inicio
    horario.hora_fin = hora_fin
    horario.actividad = actividad
    db.commit()
    return RedirectResponse(url="/admin/gestionar-horarios", status_code=HTTP_303_SEE_OTHER)

@app.get("/admin/eliminar-horario/{horario_id}")
def eliminar_horario(horario_id: int, db=Depends(get_db)):
    horario = db.query(Horario).filter(Horario.id == horario_id).first()
    if horario:
        db.delete(horario)
        db.commit()
    return RedirectResponse(url="/admin/gestionar-horarios", status_code=HTTP_303_SEE_OTHER)

@app.get("/admin/publicar-video", response_class=HTMLResponse)
def mostrar_formulario_video(request: Request):
    return templates.TemplateResponse("publicar_video.html", {"request": request})

@app.post("/admin/publicar-video")
def publicar_video(request: Request, title: str = Form(...), url: str = Form(...), plataforma: str = Form(...), db: Session = Depends(get_db)):
    try:
        embed_url = generar_embed(url)
    except Exception:
        return templates.TemplateResponse("publicar_video.html", {
            "request": request,
            "error": "URL inválida o plataforma no soportada."
        })
    nuevo_post = Post(titulo=title, url=url, plataforma=plataforma, embed_url=embed_url, publicado=True)
    db.add(nuevo_post)
    db.commit()
    return RedirectResponse(url="/admin", status_code=303)

@app.post("/admin/editar/{titulo}")
def editar_seccion_post(request: Request, titulo: str, contenido: str = Form(...), imagen_url: str = Form(None), db=Depends(get_db)):
    seccion = db.query(SeccionInformativa).filter_by(titulo=titulo).first()
    if not seccion:
        raise HTTPException(status_code=404, detail="Sección no encontrada")
    if imagen_url:
        contenido += f'<br><img src="{imagen_url}" alt="Imagen relacionada" style="max-width:100%;">'
    seccion.contenido = contenido
    db.commit()
    return RedirectResponse(url="/admin", status_code=303)



@app.get("/mision", response_class=HTMLResponse)
async def seccion_mision(request: Request, db: Session = Depends(get_db)):
    seccion = db.query(SeccionInformativa).filter_by(slug="mision").first()
    return templates.TemplateResponse("ver_seccion.html", {
        "request": request,
        "titulo": seccion.titulo if seccion else "Misión",
        "contenido": seccion.contenido if seccion else "Contenido no disponible.",
        "imagen_url": seccion.imagen_url if seccion else None,
        "secciones": db.query(SeccionInformativa).all()
    })

@app.get("/vision", response_class=HTMLResponse)
async def seccion_vision(request: Request, db: Session = Depends(get_db)):
    seccion = db.query(SeccionInformativa).filter_by(slug="vision").first()
    return templates.TemplateResponse("ver_seccion.html", {
        "request": request,
        "titulo": seccion.titulo if seccion else "Visión",
        "contenido": seccion.contenido if seccion else "Contenido no disponible.",
        "imagen_url": seccion.imagen_url if seccion else None,
        "secciones": db.query(SeccionInformativa).all()
    })

@app.get("/quienes-somos", response_class=HTMLResponse)
async def seccion_quienes(request: Request, db: Session = Depends(get_db)):
    seccion = db.query(SeccionInformativa).filter_by(slug="quienes-somos").first()
    return templates.TemplateResponse("ver_seccion.html", {
        "request": request,
        "titulo": seccion.titulo if seccion else "¿Quiénes somos?",
        "contenido": seccion.contenido if seccion else "Contenido no disponible.",
        "imagen_url": seccion.imagen_url if seccion else None,
        "secciones": db.query(SeccionInformativa).all()
    })


@app.get("/servicios", response_class=HTMLResponse)
def servicios(request: Request, db=Depends(get_db)):
    seccion = db.query(SeccionInformativa).filter(SeccionInformativa.titulo == "Nuestros Servicios").first()
    secciones = db.query(SeccionInformativa).all()  # Para el menú lateral

    return templates.TemplateResponse("servicios.html", {
        "request": request,
        "contenido": seccion,
        "secciones": secciones
    })
    

@app.get("/admin/reporte-suscriptores", response_class=HTMLResponse)
async def ver_reporte_suscriptores(
    request: Request,
    db: Session = Depends(get_db),
    admin: bool = Depends(check_admin_logged)
):
    suscriptores = db.query(Suscriptor).all()
    return templates.TemplateResponse("reporte_suscriptores.html", {
        "request": request,
        "suscriptores": suscriptores
    })
    
@app.get("/admin/descargar-suscriptores")
async def descargar_excel_suscriptores(
    db: Session = Depends(get_db),
    admin: bool = Depends(check_admin_logged)
):
    suscriptores = db.query(Suscriptor).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Suscriptores"

    # --- Insertar imagen/logo si existe
    logo_path = "app/static/logo.png"
    if os.path.exists(logo_path):
        img = ExcelImage(logo_path)
        img.width = 100
        img.height = 100
        ws.add_image(img, "A1")

    # --- Título del club
    ws.merge_cells("B2:F2")
    ws["B2"] = "Deportivo Blanca Cristiana"
    ws["B2"].font = Font(size=16, bold=True)
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center")

    # --- Subtítulo
    ws.merge_cells("B4:F4")
    ws["B4"] = "Planilla de Suscriptores"
    ws["B4"].font = Font(size=14, bold=True, color="FFFFFF")
    ws["B4"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B4"].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # --- Fecha de generación
    ws.merge_cells("B5:F5")
    fecha = datetime.datetime.now().strftime("Generado el %d/%m/%Y %H:%M:%S")
    ws["B5"] = fecha
    ws["B5"].font = Font(italic=True, color="888888")
    ws["B5"].alignment = Alignment(horizontal="center")

    # --- Encabezados
    headers = ["ID", "Tipo Documento", "Número Documento", "Nombre Completo", "Correo", "Celular"]
    ws.append([])
    ws.append([])
    ws.append(headers)
    header_row = ws.max_row

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="B7D7F7", end_color="B7D7F7", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # --- Datos
    for s in suscriptores:
        ws.append([
            s.id,
            s.tipo_documento,
            s.numero_documento,
            s.nombre_completo,
            s.correo,
            s.celular
        ])

    # Ajustar ancho de columnas automáticamente
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # Exportar Excel
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=suscriptores.xlsx"}
    )
    

@app.post("/admin/eliminar-suscriptor/{suscriptor_id}")
def eliminar_suscriptor(
    suscriptor_id: int,
    request: Request,
    db: Session = Depends(get_db),
    admin: bool = Depends(check_admin_logged)
):
    suscriptor = db.query(Suscriptor).filter(Suscriptor.id == suscriptor_id).first()
    if suscriptor:
        db.delete(suscriptor)
        db.commit()
        request.session["mensaje"] = f"✅ Suscriptor con ID {suscriptor_id} eliminado exitosamente."
    else:
        request.session["mensaje"] = f"⚠️ No se encontró el suscriptor con ID {suscriptor_id}."
    
    return RedirectResponse(url="/admin/reporte-suscriptores", status_code=303)